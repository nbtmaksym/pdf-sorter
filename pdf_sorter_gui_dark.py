import os
import customtkinter as ctk
from tkinter import filedialog, messagebox

import re
import shutil
import time
import logging
import threading


try:
    import fitz
    import pytesseract
    import numpy as np
    import cv2
    from PIL import Image
    os.environ.setdefault("OMP_THREAD_LIMIT", "1")

    if not shutil.which(pytesseract.pytesseract.tesseract_cmd):
        for candidate in (
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        ):
            if os.path.isfile(candidate):
                pytesseract.pytesseract.tesseract_cmd = candidate
                break
        else:
            raise ImportError("tesseract.exe nie znaleziony")

    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

from concurrent.futures import ThreadPoolExecutor, as_completed

PT_TO_MM = 0.3528

ROZMIARY = {
    "A0": (841, 1189),
    "A1": (594, 841),
    "A2": (420, 594),
    "A3": (297, 420),
    "A4": (210, 297),
    "A5": (148, 210),
    "A6": (105, 148),
}

FOLDER_MAPPING = {
    "A0": "A2",
    "A1": "A2",
    "A2": "A2",
    "A3": "A3",
    "A4": "A4",
    "A5": "A5",
    "A6": "A6",
    "NIEZNANY": "NIEZNANY",
}

NAZWA_PREFIKSY = [
    ("HEAA", "belki_dwuteowe", 0), ("HEM", "belki_dwuteowe", 0), ("HEB", "belki_dwuteowe", 0),
    ("HEA", "belki_dwuteowe", 0), ("IPE", "belki_dwuteowe", 0), ("IPN", "belki_dwuteowe", 0),
    ("HD ", "belki_dwuteowe", 0), ("HP ", "belki_dwuteowe", 0), ("UB", "belki_dwuteowe", 0),
    ("UC", "belki_dwuteowe", 0), ("UBP", "belki_dwuteowe", 0), ("W ", "belki_dwuteowe", 0),
    ("W1", "belki_dwuteowe", 0), ("W2", "belki_dwuteowe", 0), ("W3", "belki_dwuteowe", 0),
    ("W4", "belki_dwuteowe", 0), ("W5", "belki_dwuteowe", 0), ("W6", "belki_dwuteowe", 0),
    ("W8", "belki_dwuteowe", 0), ("S ", "belki_dwuteowe", 0),
    ("UPN", "ceowniki", 1), ("UPE", "ceowniki", 1), ("UPA", "ceowniki", 1), ("UNP", "ceowniki", 1),
    ("PFC", "ceowniki", 1), ("MC", "ceowniki", 1), ("CH", "ceowniki", 1), ("BLU", "ceowniki", 1),
    ("U", "ceowniki", 1),
    ("L ", "katowniki", 2), ("L1", "katowniki", 2), ("L2", "katowniki", 2), ("L3", "katowniki", 2),
    ("L4", "katowniki", 2), ("L5", "katowniki", 2), ("L6", "katowniki", 2), ("L7", "katowniki", 2),
    ("L8", "katowniki", 2), ("L9", "katowniki", 2), ("EA", "katowniki", 2), ("UA", "katowniki", 2),
    ("BLL", "katowniki", 2),
    ("WT", "teowniki", 0), ("MT", "teowniki", 0), ("T ", "teowniki", 0), ("T1", "teowniki", 0),
    ("T2", "teowniki", 0), ("T3", "teowniki", 0),
    ("CFRHS", "rury_prostokat", 2), ("CFSHS", "rury_prostokat", 2), ("RHS", "rury_prostokat", 2),
    ("SHS", "rury_prostokat", 2), ("HSS", "rury_prostokat", 2), ("TR", "rury_prostokat", 2),
    ("QRRC", "rury_prostokat", 2), ("HQ", "rury_prostokat", 2), ("HR", "rury_prostokat", 2),
    ("QRC", "rury_okragle", 1), ("QRO", "rury_okragle", 1),
    ("CHS", "rury_okragle", 1), ("RO", "rury_okragle", 1), ("OB", "rury_okragle", 1),
    ("SO", "rury_okragle", 1), ("SH", "rury_okragle", 1), ("E ", "rury_okragle", 1),
    ("RU", "prety_okragle", 0), ("RUND", "prety_okragle", 0), ("RND", "prety_okragle", 0), ("ROD", "prety_okragle", 0),
    ("PLT", "blachy", 0), ("PL", "blachy", 0), ("BL", "blachy", 0),
    ("FB", "blachy", 1), ("FL", "blachy", 1),
    ("ZED", "profile_specjalne", 1), ("ZETA", "profile_specjalne", 1), ("SIGMA", "profile_specjalne", 1),
    ("OMEGA", "profile_specjalne", 1), ("HAT", "profile_specjalne", 1), ("SF", "profile_specjalne", 1),
    ("Z ", "profile_specjalne", 1),
]


def _build_profile_patterns(prefixes):
    patterns = []
    digit_group = r"[\dOo]{1,3}(?![\dOo])"
    for prefix, category, extra_groups in prefixes:
        requires_space = prefix.endswith(" ")
        p = prefix.rstrip()
        sep = r"\s" if requires_space else r"\s?"
        extra = r"(?:[.,]?[xX*]" + digit_group + r")"
        pattern_str = r"\b" + re.escape(p) + sep + digit_group + extra + "{0," + str(extra_groups) + "}"
        patterns.append(re.compile(pattern_str, re.IGNORECASE))
    return patterns


PROFILE_PATTERNS = _build_profile_patterns(NAZWA_PREFIKSY)


def classify(width_mm, height_mm, tolerance=5):
    w, h = sorted([width_mm, height_mm])
    for name, (sw, sh) in ROZMIARY.items():
        if abs(w - sw) <= tolerance and abs(h - sh) <= tolerance:
            return name
    return "NIEZNANY"


STEEL_GRADE_RE = re.compile(r"[S$]\W{0,2}\d{3}\W{0,2}[A-Z0-9]{0,3}", re.IGNORECASE)
STAHLBLECH_RE = re.compile(r"STAHLBLECH\s*(\d+)\s*MM", re.IGNORECASE)


def _clean_profile_name(raw):
    cleaned = re.sub(r"\s+", "", raw)
    return cleaned.replace("*", "X")


def find_profile(text):
    text_up = text.upper()
    m = STAHLBLECH_RE.search(text_up)
    if m:
        return f"BL{m.group(1)}"

    grade_spans = [m.span() for m in STEEL_GRADE_RE.finditer(text_up)]

    candidates = []
    for pattern in PROFILE_PATTERNS:
        for m in pattern.finditer(text_up):
            candidates.append(m)

    if not candidates:
        return "NIEZNANY_PROFIL"

    def near_grade(m, window=40):
        return any(0 <= gs - m.end() <= window for gs, ge in grade_spans)

    grade_backed = [m for m in candidates if near_grade(m)]
    if grade_backed:
        def confidence(m):
            letters = re.match(r"[A-Za-z]+", m.group())
            return (len(letters.group()) if letters else 0, len(m.group()))
        best = max(grade_backed, key=confidence)
        return _clean_profile_name(best.group())

    return _clean_profile_name(candidates[0].group())


def unique_path(dest_folder, filename):
    dest = os.path.join(dest_folder, filename)
    if not os.path.exists(dest):
        return dest
    base, ext = os.path.splitext(filename)
    i = 1
    while True:
        new = os.path.join(dest_folder, f"{base}_{i}{ext}")
        if not os.path.exists(new):
            return new
        i += 1


def _safe_copy2(src, dst, retries=6, delay=0.4):
    last_err = None
    for attempt in range(retries):
        try:
            shutil.copy2(src, dst)
            return
        except (PermissionError, OSError) as e:
            last_err = e
            time.sleep(delay)
    raise last_err


def _safe_move(src, dst, retries=6, delay=0.4):
    """shutil.move z ponawianiem - patrz _safe_copy2."""
    last_err = None
    for attempt in range(retries):
        try:
            shutil.move(src, dst)
            return
        except (PermissionError, OSError) as e:
            last_err = e
            time.sleep(delay)
    raise last_err


def ocr_extract_text(filepath, target_max_px=3000, timeout_s=20):
    if not OCR_AVAILABLE:
        return ""
    try:
        doc = fitz.open(filepath)
        page = doc[0]
        page_w, page_h = page.rect.width, page.rect.height
        longer_side = max(page_w, page_h, 1)
        zoom = min(4.0, max(1.5, target_max_px / longer_side))

        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        mode = "RGB" if pix.n < 4 else "RGBA"
        img = Image.frombytes(mode, (pix.width, pix.height), pix.samples)
        if mode == "RGBA":
            img = img.convert("RGB")
        doc.close()

        gray = np.array(img.convert("L"))
        _, binarized = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        bordered = cv2.copyMakeBorder(binarized, 25, 25, 25, 25,
                                       cv2.BORDER_CONSTANT, value=255)
        proc_img = Image.fromarray(bordered)

        text = pytesseract.image_to_string(proc_img, lang="eng", config="--psm 6",
                                            timeout=timeout_s)
        if find_profile(text) == "NIEZNANY_PROFIL":
            # tryb "rozproszony tekst" jako drugie podejscie - wolniejszy,
            # wiec probowany tylko gdy pierwszy tryb nic nie znalazl
            text2 = pytesseract.image_to_string(proc_img, lang="eng", config="--psm 11",
                                                 timeout=timeout_s)
            if find_profile(text2) != "NIEZNANY_PROFIL":
                return text2
        return text
    except Exception:
        return ""


def process_folder_by_format(input_folder, log_func):
    logging.basicConfig(
        filename=os.path.join(input_folder, "pdf_sorter.log"),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        encoding="utf-8"
    )

    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pdf")]
    if not pdf_files:
        return 0, 0

    sukces, bledy = 0, 0

    wszystkie_folder = os.path.join(input_folder, "wszystkie")
    os.makedirs(wszystkie_folder, exist_ok=True)

    for filename in pdf_files:
        filepath = os.path.join(input_folder, filename)
        try:
            doc = fitz.open(filepath)
            page = doc[0]
            w_mm = page.rect.width * PT_TO_MM
            h_mm = page.rect.height * PT_TO_MM
            doc.close()
            format_name = classify(w_mm, h_mm)
            folder_name = FOLDER_MAPPING[format_name]

            # kopia do folderu "wszystkie" - PRZED przeniesieniem do
            # folderu wg formatu, zeby miec podglad calosci w jednym
            # miejscu obok normalnego sortowania
            wszystkie_dest = unique_path(wszystkie_folder, filename)
            _safe_copy2(filepath, wszystkie_dest)

            dest_folder = os.path.join(input_folder, folder_name)
            os.makedirs(dest_folder, exist_ok=True)
            dest_path = unique_path(dest_folder, filename)
            _safe_move(filepath, dest_path)

            dest_filename = os.path.basename(dest_path)
            extra = f" [wykryto: {format_name}]" if format_name != folder_name else ""
            info = f"OK  {filename} -> {folder_name}{extra} | {w_mm:.1f} x {h_mm:.1f} mm"
            if dest_filename != filename:
                info += f" (zapisano jako: {dest_filename})"
            log_func(info)
            logging.info(info)
            sukces += 1

        except Exception as e:
            msg = f"BLAD  {filename}: {e}"
            log_func(msg)
            logging.error(msg)
            bledy += 1

    return sukces, bledy


def process_folder_by_profile(input_folder, log_func, ocr_workers=4, ocr_timeout_s=20):
    logging.basicConfig(
        filename=os.path.join(input_folder, "pdf_sorter_profil.log"),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        encoding="utf-8"
    )

    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pdf")]
    if not pdf_files:
        return 0, 0

    sukces, bledy = 0, 0

    if not OCR_AVAILABLE:
        log_func("Uwaga: OCR niedostepny (brak pytesseract/fitz/Tesseract-OCR) - "
                  "pliki bez tekstu trafia do NIEZNANY_PROFIL.\n")

    # Faza 1: szybki odczyt tekstu z PDF (bez OCR) dla wszystkich plikow.
    # Zbieramy tez liste tych, ktore beda potrzebowaly OCR.
    results = {}   # filename -> dict(w_mm, h_mm, folder_name, profile_name, used_ocr)
    needs_ocr = []

    for i, filename in enumerate(pdf_files, start=1):
        filepath = os.path.join(input_folder, filename)
        try:
            doc = fitz.open(filepath)
            page = doc[0]
            w_mm = page.rect.width * PT_TO_MM
            h_mm = page.rect.height * PT_TO_MM
            format_name = classify(w_mm, h_mm)
            folder_name = FOLDER_MAPPING[format_name]
            text = page.get_text() or ""
            doc.close()
            profile_name = find_profile(text)

            results[filename] = {
                "w_mm": w_mm, "h_mm": h_mm, "folder_name": folder_name,
                "profile_name": profile_name, "used_ocr": False,
            }
            if profile_name == "NIEZNANY_PROFIL" and OCR_AVAILABLE:
                needs_ocr.append(filename)

        except Exception as e:
            msg = f"BLAD  {filename}: {e}"
            log_func(msg)
            logging.error(msg)
            bledy += 1

        if i % 50 == 0 or i == len(pdf_files):
            log_func(f"Odczyt tekstu: {i}/{len(pdf_files)}...")

    # Faza 2: OCR rownolegle (kilka plikow na raz) - to jedyna wolna
    # czesc calego procesu, wiec tylko ja jest zrownoleglona. Tesseract
    # dziala jako osobny proces systemowy, wiec ThreadPoolExecutor daje
    # tu prawdziwe przyspieszenie mimo GIL-a w Pythonie.
    if needs_ocr:
        log_func(f"OCR: przetwarzanie {len(needs_ocr)} plikow bez tekstu "
                  f"({ocr_workers} rownolegle, limit {ocr_timeout_s}s/plik)...")
        ocr_texts = {}
        done_count = 0
        with ThreadPoolExecutor(max_workers=ocr_workers) as executor:
            future_to_name = {
                executor.submit(ocr_extract_text, os.path.join(input_folder, f),
                                 timeout_s=ocr_timeout_s): f
                for f in needs_ocr
            }
            for future in as_completed(future_to_name):
                filename = future_to_name[future]
                try:
                    ocr_texts[filename] = future.result()
                except Exception:
                    ocr_texts[filename] = ""
                done_count += 1
                if done_count % 10 == 0 or done_count == len(needs_ocr):
                    log_func(f"OCR: {done_count}/{len(needs_ocr)} gotowe...")
        for filename in needs_ocr:
            ocr_profile = find_profile(ocr_texts.get(filename, ""))
            if ocr_profile != "NIEZNANY_PROFIL":
                results[filename]["profile_name"] = ocr_profile + "_OCR"
                results[filename]["used_ocr"] = True

    # Faza 3: przenoszenie plikow i log
    for filename, r in results.items():
        filepath = os.path.join(input_folder, filename)
        try:
            dest_folder = os.path.join(input_folder, r["folder_name"], r["profile_name"])
            os.makedirs(dest_folder, exist_ok=True)
            dest_path = unique_path(dest_folder, filename)
            _safe_move(filepath, dest_path)

            dest_filename = os.path.basename(dest_path)
            ocr_note = " [odczytano przez OCR - sprawdz]" if r["used_ocr"] else ""
            info = (f"OK  {filename} -> {r['folder_name']}/{r['profile_name']} | "
                    f"{r['w_mm']:.1f} x {r['h_mm']:.1f} mm{ocr_note}")
            if dest_filename != filename:
                info += f" (zapisano jako: {dest_filename})"
            log_func(info)
            logging.info(info)
            sukces += 1
        except Exception as e:
            msg = f"BLAD  {filename}: {e}"
            log_func(msg)
            logging.error(msg)
            bledy += 1

    return sukces, bledy

def process_folder_by_profile_ocr_only(input_folder, log_func, ocr_workers=4, ocr_timeout_s=20):
    if not OCR_AVAILABLE:
        log_func("BLAD: OCR niedostepny (brak pytesseract/fitz/Tesseract-OCR/opencv/numpy) - "
                  "ten tryb wymaga wszystkich tych bibliotek.")
        return 0, 0

    logging.basicConfig(
        filename=os.path.join(input_folder, "pdf_sorter_profil_ocr.log"),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        encoding="utf-8"
    )

    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pdf")]
    if not pdf_files:
        return 0, 0

    sukces, bledy = 0, 0
    sizes = {}
    for filename in pdf_files:
        filepath = os.path.join(input_folder, filename)
        try:
            doc = fitz.open(filepath)
            page = doc[0]
            w_mm = page.rect.width * PT_TO_MM
            h_mm = page.rect.height * PT_TO_MM
            doc.close()
            format_name = classify(w_mm, h_mm)
            sizes[filename] = (w_mm, h_mm, FOLDER_MAPPING[format_name])
        except Exception as e:
            msg = f"BLAD  {filename}: {e}"
            log_func(msg)
            logging.error(msg)
            bledy += 1

    ocr_files = [f for f in pdf_files if f in sizes]
    log_func(f"OCR: przetwarzanie {len(ocr_files)} plikow "
              f"({ocr_workers} rownolegle, limit {ocr_timeout_s}s/plik)...")

    ocr_texts = {}
    done_count = 0
    with ThreadPoolExecutor(max_workers=ocr_workers) as executor:
        future_to_name = {
            executor.submit(ocr_extract_text, os.path.join(input_folder, f),
                             timeout_s=ocr_timeout_s): f
            for f in ocr_files
        }
        for future in as_completed(future_to_name):
            filename = future_to_name[future]
            try:
                ocr_texts[filename] = future.result()
            except Exception:
                ocr_texts[filename] = ""
            done_count += 1
            if done_count % 10 == 0 or done_count == len(ocr_files):
                log_func(f"OCR: {done_count}/{len(ocr_files)} gotowe...")

    for filename in ocr_files:
        filepath = os.path.join(input_folder, filename)
        w_mm, h_mm, folder_name = sizes[filename]
        try:
            profile_name = find_profile(ocr_texts.get(filename, ""))
            if profile_name == "NIEZNANY_PROFIL":
                profile_folder = "NIEZNANY_PROFIL"
            else:
                profile_folder = profile_name + "_OCR"

            dest_folder = os.path.join(input_folder, folder_name, profile_folder)
            os.makedirs(dest_folder, exist_ok=True)
            dest_path = unique_path(dest_folder, filename)
            _safe_move(filepath, dest_path)

            dest_filename = os.path.basename(dest_path)
            info = (f"OK  {filename} -> {folder_name}/{profile_folder} | "
                    f"{w_mm:.1f} x {h_mm:.1f} mm")
            if dest_filename != filename:
                info += f" (zapisano jako: {dest_filename})"
            log_func(info)
            logging.info(info)
            sukces += 1
        except Exception as e:
            msg = f"BLAD  {filename}: {e}"
            log_func(msg)
            logging.error(msg)
            bledy += 1

    return sukces, bledy


try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

import difflib

ANTYKOROZJA_FOLDERY = {"M": "malowane", "C": "cynkowane", "D": "duplex"}


QTY_TAIL_RE = re.compile(r"\s+\d{1,3}\W{0,2}[xX]\s*$")


def load_wykaz_antykorozja(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.strip().lower()
        if "wykaz" not in name_lower or "spoin" not in name_lower:
            continue
        ws = wb[sheet_name]
        header_row_idx, col_pozycja, col_antykorozja = None, None, None
        for r in range(1, min(20, ws.max_row) + 1):
            for c in range(1, min(60, ws.max_column) + 1):
                val = ws.cell(row=r, column=c).value
                if not val:
                    continue
                v = str(val).strip().lower().replace("\n", " ")
                if "pozycja" in v:
                    col_pozycja, header_row_idx = c, r
                if "antykorozja" in v:
                    col_antykorozja, header_row_idx = c, r
            if col_pozycja and col_antykorozja:
                break
        if not (col_pozycja and col_antykorozja and header_row_idx):
            continue
        for r in range(header_row_idx + 1, ws.max_row + 1):
            pos_val = ws.cell(row=r, column=col_pozycja).value
            corr_val = ws.cell(row=r, column=col_antykorozja).value
            if pos_val is None:
                continue
            pos_key = str(pos_val).strip()
            corr_key = str(corr_val).strip().upper() if corr_val is not None else ""
            if pos_key:
                result[pos_key] = corr_key
    return result


def extract_pos_chunk(text):
    m = re.search(r"POS\.?\s*([^\n]{1,25})", text, re.IGNORECASE)
    if not m:
        return None
    chunk = QTY_TAIL_RE.sub("", m.group(1))
    chunk = chunk.strip(" -+–—?.")
    return chunk or None


def _normalize_pos_token(token, keep_unknown=False):
    repl = {"O": "0", "I": "1", "L": "1", "S": "5", "B": "8", " ": "_"}
    out = []
    for ch in token.upper():
        if ch.isdigit() or ch == "_":
            out.append(ch)
        elif ch in repl:
            out.append(repl[ch])
        elif keep_unknown:
            out.append(ch)
    return "".join(out)


def lookup_antykorozja(text, wykaz, fuzzy_cutoff=0.75, ambiguity_margin=0.05, wykaz_norm=None):
    chunk = extract_pos_chunk(text)
    if not chunk:
        return None, None, None, "brak-tekstu-Pos"

    raw_candidates = {chunk}
    if " " in chunk:
        raw_candidates.add(chunk.replace(" ", "_"))

    for c in raw_candidates:
        if c in wykaz:
            return wykaz[c], c, 1.0, "dokladne"

    if wykaz_norm is None:
        wykaz_norm = {_normalize_pos_token(k, keep_unknown=True): k for k in wykaz}
    for c in raw_candidates:
        n = _normalize_pos_token(c, keep_unknown=True)
        if n in wykaz_norm:
            key = wykaz_norm[n]
            return wykaz[key], key, 0.95, "znormalizowane"

    best_query = _normalize_pos_token(max(raw_candidates, key=len), keep_unknown=True)
    scored = sorted(
        ((difflib.SequenceMatcher(None, best_query, _normalize_pos_token(k, keep_unknown=True)).ratio(), k)
         for k in wykaz),
        reverse=True
    )
    if scored and scored[0][0] >= fuzzy_cutoff:
        if len(scored) == 1 or (scored[0][0] - scored[1][0]) >= ambiguity_margin:
            score, key = scored[0]
            return wykaz[key], key, round(score, 2), "rozmyte"
        return None, best_query, round(scored[0][0], 2), "wieloznaczne"

    return None, best_query, None, "brak-dopasowania"


def process_folder_by_corrosion(input_folder, log_func, wykaz_path,
                                 ocr_workers=4, ocr_timeout_s=20):
    if not OCR_AVAILABLE:
        log_func("BLAD: OCR niedostepny - ten tryb wymaga pytesseract/fitz/opencv/numpy.")
        return 0, 0
    if not EXCEL_AVAILABLE:
        log_func("BLAD: brak biblioteki openpyxl - zainstaluj: pip install openpyxl")
        return 0, 0
    if not wykaz_path or not os.path.exists(wykaz_path):
        log_func("BLAD: nie wybrano poprawnego pliku wykazu wysylkowego (.xlsx).")
        return 0, 0

    try:
        wykaz = load_wykaz_antykorozja(wykaz_path)
    except Exception as e:
        log_func(f"BLAD: nie udalo sie odczytac wykazu: {e}")
        return 0, 0

    if not wykaz:
        log_func("BLAD: w wykazie nie znaleziono kolumn 'Pozycja'/'Antykorozja' "
                  "(sprawdz czy arkusz nazywa sie 'Wykaz bez spoin' / 'wykaz z spoinami').")
        return 0, 0

    log_func(f"Wczytano wykaz: {len(wykaz)} pozycji.\n")
    wykaz_norm_precomputed = {_normalize_pos_token(k, keep_unknown=True): k for k in wykaz}

    logging.basicConfig(
        filename=os.path.join(input_folder, "pdf_sorter_antykorozja.log"),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        encoding="utf-8"
    )

    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pdf")]
    if not pdf_files:
        return 0, 0

    sukces, bledy = 0, 0

    sizes = {}
    for filename in pdf_files:
        filepath = os.path.join(input_folder, filename)
        try:
            doc = fitz.open(filepath)
            page = doc[0]
            w_mm = page.rect.width * PT_TO_MM
            h_mm = page.rect.height * PT_TO_MM
            doc.close()
            format_name = classify(w_mm, h_mm)
            sizes[filename] = (w_mm, h_mm, FOLDER_MAPPING[format_name])
        except Exception as e:
            msg = f"BLAD  {filename}: {e}"
            log_func(msg)
            logging.error(msg)
            bledy += 1

    ocr_files = [f for f in pdf_files if f in sizes]
    log_func(f"OCR: przetwarzanie {len(ocr_files)} plikow "
              f"({ocr_workers} rownolegle, limit {ocr_timeout_s}s/plik)...")

    ocr_texts = {}
    done_count = 0
    with ThreadPoolExecutor(max_workers=ocr_workers) as executor:
        future_to_name = {
            executor.submit(ocr_extract_text, os.path.join(input_folder, f),
                             timeout_s=ocr_timeout_s): f
            for f in ocr_files
        }
        for future in as_completed(future_to_name):
            filename = future_to_name[future]
            try:
                ocr_texts[filename] = future.result()
            except Exception:
                ocr_texts[filename] = ""
            done_count += 1
            if done_count % 10 == 0 or done_count == len(ocr_files):
                log_func(f"OCR: {done_count}/{len(ocr_files)} gotowe...")

    for filename in ocr_files:
        filepath = os.path.join(input_folder, filename)
        w_mm, h_mm, folder_name = sizes[filename]
        try:
            text = ocr_texts.get(filename, "")
            profile_name = find_profile(text)
            profile_folder = (profile_name + "_OCR") if profile_name != "NIEZNANY_PROFIL" \
                else "NIEZNANY_PROFIL"

            kod, matched_key, confidence, metoda = lookup_antykorozja(
                text, wykaz, wykaz_norm=wykaz_norm_precomputed)
            corrosion_folder = ANTYKOROZJA_FOLDERY.get(kod, "NIEZNANE_ANTYKOROZJA")

            dest_folder = os.path.join(input_folder, corrosion_folder, folder_name, profile_folder)
            os.makedirs(dest_folder, exist_ok=True)
            dest_path = unique_path(dest_folder, filename)
            _safe_move(filepath, dest_path)

            dest_filename = os.path.basename(dest_path)
            if matched_key and kod:
                pos_note = f" [pozycja: {matched_key}, {metoda}, pewnosc={confidence}]"
            elif matched_key and metoda == "wieloznaczne":
                pos_note = (f" [odczytano '{matched_key}' - kilka podobnych pozycji w wykazie, "
                             f"nie da sie jednoznacznie wybrac (pewnosc={confidence})]")
            elif matched_key:
                pos_note = f" [odczytano '{matched_key}' - brak w wykazie]"
            else:
                pos_note = " [nie udalo sie odczytac numeru pozycji]"
            info = (f"OK  {filename} -> {corrosion_folder}/{folder_name}/{profile_folder} | "
                    f"{w_mm:.1f} x {h_mm:.1f} mm{pos_note}")
            if dest_filename != filename:
                info += f" (zapisano jako: {dest_filename})"
            log_func(info)
            logging.info(info)
            sukces += 1
        except Exception as e:
            msg = f"BLAD  {filename}: {e}"
            log_func(msg)
            logging.error(msg)
            bledy += 1

    return sukces, bledy


# --- GUI ---

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

COLOR_BG = "#1a1a1e"
COLOR_CARD = "#242429"
COLOR_ACCENT = "#3b82f6"
COLOR_ACCENT_HOVER = "#2563eb"
COLOR_TEXT_MUTED = "#8a8a94"
COLOR_LOG_BG = "#111114"
COLOR_OK = "#22c55e"
COLOR_WARN = "#f59e0b"
COLOR_ERR = "#ef4444"

FONT_TITLE = ("Segoe UI", 22, "bold")
FONT_SUBTITLE = ("Segoe UI", 12)
FONT_LABEL = ("Segoe UI", 13, "bold")
FONT_BODY = ("Segoe UI", 12)
FONT_BUTTON = ("Segoe UI", 13, "bold")
FONT_LOG = ("Consolas", 11)
FONT_STATUS = ("Segoe UI", 11)


class Tab:
    """Jedna zakladka: folder + opis + przycisk startu + log + pasek postepu."""

    def __init__(self, parent, icon, title, description, process_func,
                 button_label="Start", extra_note=None, needs_wykaz=False):
        self.process_func = process_func
        self.button_label = button_label
        self.needs_wykaz = needs_wykaz

        root = ctk.CTkFrame(parent, fg_color="transparent")
        root.pack(fill="both", expand=True, padx=20, pady=20)

        header = ctk.CTkFrame(root, fg_color="transparent")
        header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(header, text=f"{icon}  {title}", font=FONT_LABEL,
                     text_color="#e5e5ea").pack(anchor="w")

        desc_label = ctk.CTkLabel(root, text=description, font=FONT_BODY,
                                   text_color=COLOR_TEXT_MUTED, justify="left",
                                   anchor="w")
        desc_label.pack(anchor="w", fill="x", pady=(0, 14))

        note = None
        if extra_note:
            note = ctk.CTkLabel(root, text=extra_note, font=FONT_BODY,
                                 text_color=COLOR_WARN, justify="left", anchor="w")
            note.pack(anchor="w", fill="x", pady=(0, 10))
        last_width = {"value": -1}

        def _on_root_resize(event):
            if event.width == last_width["value"]:
                return
            last_width["value"] = event.width
            wl = max(200, event.width - 8)
            desc_label.configure(wraplength=wl)
            if note is not None:
                note.configure(wraplength=wl)

        root.bind("<Configure>", _on_root_resize)

        # --- karta wyboru folderu ---
        card = ctk.CTkFrame(root, fg_color=COLOR_CARD, corner_radius=12)
        card.pack(fill="x", pady=(0, 16))
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=16, pady=16)

        ctk.CTkLabel(inner, text="📂  Folder z plikami PDF", font=FONT_LABEL,
                     text_color="#e5e5ea").pack(anchor="w", pady=(0, 8))

        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x")
        self.folder_var = ctk.StringVar()
        entry = ctk.CTkEntry(row, textvariable=self.folder_var,
                              placeholder_text="Wybierz folder...",
                              height=38, font=FONT_BODY, corner_radius=8)
        entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ctk.CTkButton(row, text="Przeglądaj", width=120, height=38,
                      corner_radius=8, font=FONT_BUTTON,
                      fg_color="#3a3a42", hover_color="#48484f",
                      command=self.select_folder).pack(side="left")

        self.wykaz_var = None
        if self.needs_wykaz:
            ctk.CTkLabel(inner, text="📊  Wykaz wysyłkowy (.xlsx)", font=FONT_LABEL,
                         text_color="#e5e5ea").pack(anchor="w", pady=(14, 8))
            row2 = ctk.CTkFrame(inner, fg_color="transparent")
            row2.pack(fill="x")
            self.wykaz_var = ctk.StringVar()
            entry2 = ctk.CTkEntry(row2, textvariable=self.wykaz_var,
                                   placeholder_text="Wybierz plik wykazu wysyłkowego...",
                                   height=38, font=FONT_BODY, corner_radius=8)
            entry2.pack(side="left", fill="x", expand=True, padx=(0, 10))
            ctk.CTkButton(row2, text="Przeglądaj", width=120, height=38,
                          corner_radius=8, font=FONT_BUTTON,
                          fg_color="#3a3a42", hover_color="#48484f",
                          command=self.select_wykaz).pack(side="left")

        # --- przycisk startu ---
        self.btn_run = ctk.CTkButton(
            root, text=f"▶  {button_label}", height=46, corner_radius=10,
            font=FONT_BUTTON, fg_color=COLOR_ACCENT, hover_color=COLOR_ACCENT_HOVER,
            command=self.run
        )
        self.btn_run.pack(fill="x", pady=(0, 14))

        # --- pasek postepu ---
        self.progress = ctk.CTkProgressBar(root, height=6, corner_radius=3,
                                            progress_color=COLOR_ACCENT)
        self.progress.set(0)
        self.progress.pack(fill="x", pady=(0, 14))

        # --- log ---
        ctk.CTkLabel(root, text="Log", font=FONT_LABEL,
                     text_color="#e5e5ea").pack(anchor="w", pady=(0, 6))
        self.log_box = ctk.CTkTextbox(root, fg_color=COLOR_LOG_BG,
                                       font=FONT_LOG, corner_radius=10,
                                       text_color="#c8c8d0", wrap="none")
        self.log_box.pack(fill="both", expand=True)
        self.log_box.configure(state="disabled")

        self.running = False

    def select_folder(self):
        folder = filedialog.askdirectory(title="Wybierz folder z plikami PDF")
        if folder:
            self.folder_var.set(folder)

    def select_wykaz(self):
        path = filedialog.askopenfilename(title="Wybierz wykaz wysyłkowy",
                                           filetypes=[("Pliki Excel", "*.xlsx")])
        if path:
            self.wykaz_var.set(path)

    def log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _animate_progress(self):
        if not self.running:
            return
        current = self.progress.get()
        nxt = 0.08 if current >= 0.9 else current + 0.03
        self.progress.set(nxt)
        self.progress.after(120, self._animate_progress)

    def run(self):
        if self.running:
            return
        folder = self.folder_var.get().strip()
        if not folder or not os.path.exists(folder):
            messagebox.showerror("Błąd", "Wybierz poprawny folder!")
            return

        wykaz_path = None
        if self.needs_wykaz:
            wykaz_path = self.wykaz_var.get().strip()
            if not wykaz_path or not os.path.exists(wykaz_path):
                messagebox.showerror("Błąd", "Wybierz poprawny plik wykazu wysyłkowego (.xlsx)!")
                return

        pdf_files = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]
        if not pdf_files:
            messagebox.showwarning("Brak plików", "W wybranym folderze nie ma żadnych plików PDF.")
            return

        self.running = True
        self.btn_run.configure(state="disabled", text="⏳  Przetwarzanie...")
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress.set(0.05)
        self._animate_progress()

        self.log(f"Znaleziono {len(pdf_files)} plików PDF.\n")

        def task():
            if self.needs_wykaz:
                sukces, bledy = self.process_func(folder, self.log, wykaz_path=wykaz_path)
            else:
                sukces, bledy = self.process_func(folder, self.log)
            self.running = False
            self.progress.set(1.0)
            self.btn_run.configure(state="normal", text=f"▶  {self.button_label}")
            if bledy == 0:
                messagebox.showinfo("Sukces", f"Gotowe!\n\nPrzeniesiono: {sukces} plików\nBłędy: {bledy}")
            else:
                messagebox.showwarning(
                    "Zakończono z błędami",
                    f"Przeniesiono: {sukces} plików\nBłędy: {bledy}\n\nSzczegóły w logu powyżej."
                )

        threading.Thread(target=task, daemon=True).start()


def main():
    app = ctk.CTk()
    app.title("PDF Sorter")
    app.geometry("1000x760")
    app.configure(fg_color=COLOR_BG)
    app.minsize(760, 560)

    header = ctk.CTkFrame(app, fg_color="transparent")
    header.pack(fill="x", padx=24, pady=(20, 4))
    ctk.CTkLabel(header, text="📐  PDF Sorter", font=FONT_TITLE,
                 text_color="#ffffff").pack(anchor="w")
    ctk.CTkLabel(header, text="Elmo S.A. — sortowanie rysunkow PDF wg formatu i profilu",
                 font=FONT_SUBTITLE, text_color=COLOR_TEXT_MUTED).pack(anchor="w", pady=(2, 0))

    tabview = ctk.CTkTabview(app, fg_color=COLOR_BG, segmented_button_fg_color=COLOR_CARD,
                              segmented_button_selected_color=COLOR_ACCENT,
                              segmented_button_selected_hover_color=COLOR_ACCENT_HOVER,
                              segmented_button_unselected_color=COLOR_CARD,
                              text_color="#e5e5ea", corner_radius=12)
    tabview.pack(fill="both", expand=True, padx=24, pady=(10, 10))

    tab1 = tabview.add("📁  Wg formatu")
    tab2 = tabview.add("🔍  Wg formatu + profilu")
    tab3 = tabview.add("🔬  Tylko OCR")
    tab4 = tabview.add("🛡️  Antykorozja + profil")

    Tab(tab1, "📁", "Sortowanie wg formatu strony",
        "Klasyfikuje pliki PDF wedlug wielkosci strony (A2, A3, A4...) i "
        "przenosi je do odpowiednich podfolderow. Formaty A0 i A1 traktowane "
        "sa jako A2.",
        process_folder_by_format)

    Tab(tab2, "🔍", "Sortowanie wg formatu i profilu",
        "Najpierw sortuje wedlug formatu strony, a wewnatrz kazdego formatu "
        "tworzy podfoldery wedlug wykrytego profilu stalowego (np. HEA200, "
        "PL5, RHS100x50x5). Dla plikow bez warstwy tekstowej automatycznie "
        "probuje odczytac profil przez OCR.",
        process_folder_by_profile,
        extra_note="⚠  Dla plikow bez tekstu wymaga zainstalowanego Tesseract-OCR."
        if not OCR_AVAILABLE else None)

    Tab(tab3, "🔬", "Sortowanie tylko przez OCR",
        "Pomija probe odczytu prawdziwego tekstu z PDF i od razu uzywa OCR "
        "dla kazdego pliku. Uzyj tego trybu jesli wiesz z gory, ze wszystkie "
        "pliki w folderze nie maja warstwy tekstowej (np. detale wyciete z "
        "rysunkow Tekla).",
        process_folder_by_profile_ocr_only,
        extra_note="⚠  Wymaga zainstalowanego Tesseract-OCR." if not OCR_AVAILABLE else None)

    corrosion_note = None
    if not OCR_AVAILABLE:
        corrosion_note = "⚠  Wymaga zainstalowanego Tesseract-OCR."
    elif not EXCEL_AVAILABLE:
        corrosion_note = "⚠  Wymaga biblioteki openpyxl (pip install openpyxl)."

    Tab(tab4, "🛡️", "Sortowanie wg antykorozji i profilu",
        "Dopasowuje kazdy detal do wykazu wysylkowego po numerze pozycji "
        "(odczytanym przez OCR z etykiety 'Pos. XXX') i najpierw dzieli na "
        "malowane / cynkowane / duplex wedlug kolumny 'Antykorozja', a "
        "dopiero w srodku sortuje wg formatu i profilu jak w zakladce OCR. "
        "Wykaz musi miec arkusz 'Wykaz bez spoin' lub 'wykaz z spoinami' z "
        "kolumnami 'Pozycja' i 'Antykorozja'.",
        process_folder_by_corrosion,
        extra_note=corrosion_note,
        needs_wykaz=True)

    status = ctk.CTkLabel(
        app,
        text=("OCR: dostepny ✓" if OCR_AVAILABLE else "OCR: niedostepny — zainstaluj Tesseract-OCR i pytesseract"),
        font=FONT_STATUS,
        text_color=COLOR_OK if OCR_AVAILABLE else COLOR_WARN,
    )
    status.pack(anchor="w", padx=28, pady=(0, 14))

    app.mainloop()


if __name__ == "__main__":
    main()